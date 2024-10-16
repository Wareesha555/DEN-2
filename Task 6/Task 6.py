import openpyxl

TASK_FILE = 'tasks.xlsx'  # Excel file will be saved in the current working directory

def load_tasks():
    """Load tasks from the Excel file."""
    try:
        workbook = openpyxl.load_workbook(TASK_FILE)
        sheet = workbook.active
        tasks = []

        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
            task = {
                'id': row[0],
                'description': row[1],
                'status': row[2]
            }
            tasks.append(task)
        return tasks
    except FileNotFoundError:
        return []  # Return an empty list if the file does not exist

def save_tasks(tasks):
    """Save tasks to the Excel file."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Tasks"

    # Write header
    sheet.append(['ID', 'Description', 'Status'])

    # Write tasks
    for task in tasks:
        sheet.append([task['id'], task['description'], task['status']])

    workbook.save(TASK_FILE)  # Save the workbook

def generate_task_id(tasks):
    """Generate a unique task ID."""
    if tasks:
        return max(task['id'] for task in tasks) + 1  # Increment max ID by 1
    else:
        return 1  # Start from ID 1 if no tasks exist

def add_task(tasks):
    """Add a new task to the list."""
    description = input("Enter the task description: ")
    task = {
        'id': generate_task_id(tasks),  # Generate unique ID
        'description': description,
        'status': 'pending'  # Default status
    }
    tasks.append(task)  # Add task to the list
    print(f"Task added with ID: {task['id']}")

def view_tasks(tasks):
    """View all tasks."""
    if not tasks:
        print("No tasks found.")
    else:
        for task in tasks:
            print(f"ID: {task['id']}, Description: {task['description']}, Status: {task['status']}")

def remove_task(tasks):
    """Remove a task by ID."""
    task_id = int(input("Enter the task ID to remove: "))
    task_to_remove = next((task for task in tasks if task['id'] == task_id), None)

    if task_to_remove:
        tasks.remove(task_to_remove)  # Remove the task
        print(f"Task with ID {task_id} removed.")
    else:
        print(f"No task found with ID {task_id}.")

def mark_completed(tasks):
    """Mark a task as completed."""
    task_id = int(input("Enter the task ID to mark as completed: "))
    task_to_complete = next((task for task in tasks if task['id'] == task_id), None)

    if task_to_complete:
        task_to_complete['status'] = 'completed'  # Update status
        print(f"Task with ID {task_id} marked as completed.")
    else:
        print(f"No task found with ID {task_id}.")

def edit_task(tasks):
    """Edit the description of an existing task."""
    task_id = int(input("Enter the task ID to edit: "))
    task_to_edit = next((task for task in tasks if task['id'] == task_id), None)

    if task_to_edit:
        new_description = input("Enter the new description: ")
        task_to_edit['description'] = new_description  # Update description
        print(f"Task with ID {task_id} updated.")
    else:
        print(f"No task found with ID {task_id}.")

def search_task(tasks):
    """Search for tasks by keyword in the description."""
    keyword = input("Enter the keyword to search: ").lower()
    results = [task for task in tasks if keyword in task['description'].lower()]

    if results:
        for task in results:
            print(f"ID: {task['id']}, Description: {task['description']}, Status: {task['status']}")
    else:
        print("No tasks found with the given keyword.")

def filter_tasks(tasks):
    """Filter tasks based on their status."""
    status = input("Enter the status to filter by (pending/completed): ").lower()
    filtered_tasks = [task for task in tasks if task['status'] == status]

    if filtered_tasks:
        for task in filtered_tasks:
            print(f"ID: {task['id']}, Description: {task['description']}, Status: {task['status']}")
    else:
        print(f"No {status} tasks found.")

def clear_all_tasks(tasks):
    """Clear all tasks after confirmation."""
    confirm = input("Are you sure you want to clear all tasks? (yes/no): ").lower()
    if confirm == 'yes':
        tasks.clear()  # Clear the task list
        print("All tasks cleared.")
    else:
        print("Operation canceled.")

def sort_tasks(tasks):
    """Sort tasks by ID or status."""
    sort_by = input("Sort by 'id' or 'status': ").lower()

    if sort_by == 'id':
        tasks.sort(key=lambda task: task['id'])  # Sort by ID
        print("Tasks sorted by ID.")
    elif sort_by == 'status':
        tasks.sort(key=lambda task: task['status'])  # Sort by status
        print("Tasks sorted by status.")
    else:
        print("Invalid sorting option.")

def display_menu():
    """Display the main menu."""
    print("\n--- To-Do List Menu ---")
    print("1. Add Task")
    print("2. View Tasks")
    print("3. Remove Task")
    print("4. Mark as Completed")
    print("5. Edit Task")
    print("6. Search Task")
    print("7. Filter Tasks")
    print("8. Clear All Tasks")
    print("9. Sort Tasks")
    print("0. Exit")

def main():
    """Main function to run the application."""
    tasks = load_tasks()  # Load existing tasks

    while True:
        display_menu()  # Display menu options
        choice = input("Choose an option: ")

        if choice == '1':
            add_task(tasks)
        elif choice == '2':
            view_tasks(tasks)
        elif choice == '3':
            remove_task(tasks)
        elif choice == '4':
            mark_completed(tasks)
        elif choice == '5':
            edit_task(tasks)
        elif choice == '6':
            search_task(tasks)
        elif choice == '7':
            filter_tasks(tasks)
        elif choice == '8':
            clear_all_tasks(tasks)
        elif choice == '9':
            sort_tasks(tasks)
        elif choice == '0':
            save_tasks(tasks)  # Save tasks before exiting
            print("Exiting application. Goodbye!")
            break
        else:
            print("Invalid option. Please try again.")

if __name__ == '__main__':
    main()
